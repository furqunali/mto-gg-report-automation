#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MTO & G&G - FMO monthly report updater   (v1 - CORE, safe)
==========================================================
This automates the tedious + error-prone parts of your monthly process WITHOUT
touching the delicate formula structure of the Summary sheet.

WHY it does not auto-insert new items with a script:
  The "Summary Comparision Report" tab contains a reconciliation block (rows ~390)
  and an Excel structured Table ("Table1", rows ~400) whose ~146 formulas point at
  the item sub-total rows BY ABSOLUTE ROW NUMBER (=G274, =G11, SUBTOTAL(Table1[..])).
  Excel repairs those automatically when YOU insert a row in-app; a Python library
  cannot - it would silently corrupt the pivot/reconciliation.  So new items are
  detected, placed to the exact category, and handed to you to paste (Excel then
  fixes Table1 for you).  Everything else is fully automated.

WHAT it does, every month, automatically:
  1. Reads the new "Product Group Ranking" file (all 4 stores stacked in one sheet).
  2. Rewrites the 4 store sheets (0008 Northgate / 0028 Riverside / 0062 Lakeview / 0025 Summit)
     with this month's sold items  ->  the Summary's existing VLOOKUPs then pull the
     sales automatically when you open the file in Excel (let it recalculate).
  3. (optional) Widens the 3 capped VLOOKUP ranges to full-column so a longer month
     can never silently drop items.  Same cells, no rows moved -> 100% safe.
  4. Adds a "Change Log" sheet listing, for review:
        - NEW items (sold but not yet in the Summary) + their exact target category
          and per-store qty/$, so you paste them in the right place with no hunting;
        - MISSING items (in the Summary but no sales this month) = No-Movement list;
        - any category it could not map, and any unknown store name.

The join key across everything is the UPC.

------------------------------------------------------------------------------
RUN each month:  edit the CONFIG block, then:   python update_report.py
------------------------------------------------------------------------------
"""

import os
import glob
import re
from collections import OrderedDict
import openpyxl

# =============================== CONFIG =====================================
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ----------------------------------------------------------------------------
# EACH MONTH you only do TWO things:
#   1. Save the new "Product Group Ranking" you download from Focal Point into
#      the report folder (same folder as the workbooks). Keep the words
#      "Product Group Ranking" in the file name. The script auto-picks the
#      NEWEST such file - no editing needed.
#   2. Point TEMPLATE_FILE at LAST month's finished workbook (one line below).
# Then run  run_report.bat  (double-click) or  python update_report.py .
# The new month's name, output file name and dashboard title are all worked
# out automatically from the template. Leave the AUTO_* values as None.
# ----------------------------------------------------------------------------

# LAST month's completed workbook (the ONE line to update each month):
TEMPLATE_FILE = os.path.join(BASE, "01 - MTO & G&G - FMO June 2026.xlsx")

# Leave these None to auto-detect / auto-derive (recommended). Set a value to
# override for a one-off run.
RAW_FILE            = None   # None -> newest "*Product Group Ranking*.xlsx"
NEW_MONTH_LABEL     = None   # None -> template's newest month + 1 (e.g. "JULY'26")
REPORT_TITLE_PERIOD = None   # None -> e.g. "July 2026"
OUT_FILE            = None   # None -> "01 - MTO & G&G - FMO <Month> <Year>.xlsx"

FIX_VLOOKUP_RANGES = True   # widen 0008/0028/0062 capped ranges to full column
DRY_RUN            = False   # True = only print the report, write nothing

# --- Feature switches -------------------------------------------------------
UPDATE_TREND        = True   # slide the 12-month Trend window + freeze last month
ADD_STORE_TOTALS    = True   # bold TOTAL row (Qty + $) at the foot of each store tab
CLEAN_JUNK_SHEETS   = True   # drop empty Sheet1 + the bloated 'No Movement old'
# Summary-tab charts: keep the original template look (teal fill, red line,
# category titles). Done in openpyxl by restoring the title + plot-area fill
# that the library would otherwise drop - no fragile clipboard copy needed.
PRESERVE_SUMMARY_CHARTS = True
REBUILD_SUMMARY_CHARTS  = False  # alt: replace with clean blue sparklines (off by default)
BUILD_DASHBOARD     = True   # (re)build the Dashboard summary sheet + charts
BUILD_HTML_DASHBOARD = True  # also emit a standalone interactive .html dashboard
AUTO_INSERT_ITEMS   = True   # let Excel (COM) insert NEW items into the Summary
RECALC_WITH_EXCEL   = True   # let Excel recalc + save the finished file (no F9)
# ============================================================================

# Dashboard palette (categorical, CVD-validated delta-E 9.2): the three series
# match the user's concept: Total DMK = blue, Uber & GrubHub = gold, Compound = green.
CLR_LSK   = "2E5FAC"   # blue
CLR_UBER  = "E0A200"   # gold
CLR_COMPD = "2E9E5B"   # green
CLR_MTO   = "2E5FAC"   # blue   (MTO)
CLR_GG    = "6FB7E0"   # light blue (G&G) — same-hue family, distinct value
CLR_TREND = "2E9E5B"   # green line
CLR_INK   = "1F2933"   # near-black text
CLR_MUTE  = "6B7785"   # muted text
CLR_BAND  = "F2F5F9"   # tile background

# Store name as it appears in the raw file (column C)  ->  store-sheet code.
STORE_TO_SHEET = {
    "Demo Market Northgate": "0008",
    "Riverside Station":       "0028",
    "Lakeview Station":     "0062",
    "Summit Station":       "0025",
}
SHEET_DISPLAY = {v: k for k, v in STORE_TO_SHEET.items()}

# Summary per-site column pairs (qty col, $ col), 1-based, for reference.
SITE_COLS = OrderedDict([("0008", (7, 8)), ("0028", (9, 10)),
                         ("0062", (11, 12)), ("0025", (13, 14))])

# Raw category name -> Summary (left-block) category name. Only renames listed.
CAT_MAP = {
    "Breakfast Bagel (Biscuits)": "Breakfast Bagel",
    "Fried Chicken":              "Chicken Tenders",
    "DMK - Catering":             "Catering",
}

SUMMARY_SHEET  = "Summary Comparision Report"
FIRST_LEFT_ROW = 7
SECONDARY_TABLE_START = 396   # the reconciliation/Table1 area begins here; never read past it

# Trend table geometry (Summary sheet). 12 month columns T..AE, one per month;
# AE is always the newest month and is a LIVE formula pointing at the current
# Summary subtotals. Category rows 6-33, header row 5, column totals row 34.
TREND_HDR_ROW   = 5
TREND_FIRST_ROW = 6
TREND_LAST_ROW  = 33
TREND_TOTAL_ROW = 34
TREND_FIRST_COL = 20          # T = oldest month in the window
TREND_LAST_COL  = 31          # AE = newest month (live formula, do not overwrite)

MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]
_MONTH_IDX = {m[:3].lower(): i for i, m in enumerate(MONTHS)}


# ------------------------- monthly auto-resolution --------------------------
def _parse_month_label(txt):
    """Parse a Trend header like "JUNE'26" / "Jul'25" / "March'26" -> (year, m0)."""
    if not txt:
        return None
    m = re.match(r"\s*([A-Za-z]+)\s*'?\s*(\d{2,4})", str(txt))
    if not m:
        return None
    mon = _MONTH_IDX.get(m.group(1)[:3].lower())
    if mon is None:
        return None
    yy = int(m.group(2))
    return (yy + 2000 if yy < 100 else yy), mon


def resolve_config():
    """Fill in RAW_FILE / NEW_MONTH_LABEL / REPORT_TITLE_PERIOD / OUT_FILE when
    they are left None, so the only thing the user maintains is TEMPLATE_FILE
    plus dropping the ranking file into the folder. Returns a dict of paths."""
    # 1) RAW: newest "*Product Group Ranking*.xlsx" in the report folder.
    raw = RAW_FILE
    if raw is None:
        cands = [p for p in glob.glob(os.path.join(BASE, "*Product Group Ranking*.xlsx"))
                 if not os.path.basename(p).startswith("~$")]
        if not cands:
            raise SystemExit("No 'Product Group Ranking*.xlsx' found in:\n  %s\n"
                             "Save the file you downloaded from Focal Point there." % BASE)
        raw = max(cands, key=os.path.getmtime)

    # 2) New month = template's newest Trend month + 1.
    label, period = NEW_MONTH_LABEL, REPORT_TITLE_PERIOD
    if label is None or period is None:
        wb = openpyxl.load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
        hdr = wb[SUMMARY_SHEET].cell(row=TREND_HDR_ROW, column=TREND_LAST_COL).value
        wb.close()
        parsed = _parse_month_label(hdr)
        if not parsed:
            raise SystemExit("Could not read the template's newest Trend month "
                             "(cell AE5 = %r). Set NEW_MONTH_LABEL manually." % hdr)
        year, m0 = parsed
        year, m0 = (year + 1, 0) if m0 == 11 else (year, m0 + 1)
        if label is None:
            label = "%s'%02d" % (MONTHS[m0].upper(), year % 100)
        if period is None:
            period = "%s %d" % (MONTHS[m0], year)

    # 3) Output name from the period, unless overridden.
    out = OUT_FILE or os.path.join(BASE, "01 - MTO & G&G - FMO %s.xlsx" % period)
    return dict(raw=raw, template=TEMPLATE_FILE, out=out, label=label, period=period)


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _is_upc(v):
    """Real UPCs are long text ('000049999953468') or start with 'ENT'; the
    embedded pivot uses tiny integers we must ignore."""
    if v is None:
        return False
    s = str(v).strip()
    return s.upper().startswith("ENT") or (s.isdigit() and len(s) >= 10) or len(s) >= 11


# --------------------------- raw file parsing -------------------------------
def parse_raw(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    by_store = OrderedDict((c, []) for c in STORE_TO_SHEET.values())
    upc_info = {}          # upc -> (raw_category, name, unit)
    per_store = {}         # upc -> {code: (qty,$)}
    unknown_stores = set()
    store_code = cat = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        def g(i): return row[i] if len(row) > i else None
        B, C, D = g(1), g(2), g(3)
        name, upc, unit, qty, sales = g(5), g(6), g(7), g(8), g(9)
        if C and str(C).strip():
            nm = str(C).strip()
            store_code = STORE_TO_SHEET.get(nm)
            if store_code is None:
                unknown_stores.add(nm)
            cat = None
            continue
        if B and str(B).strip().endswith("Total"):
            continue
        if D and str(D).strip():
            cat = str(D).strip()
            continue
        if upc and name and store_code:
            u = str(upc).strip()
            by_store[store_code].append(dict(
                cat=cat, name=str(name).strip(), upc=u,
                unit="" if unit is None else str(unit).strip(),
                qty=_num(qty), sales=_num(sales)))
            upc_info.setdefault(u, (cat, str(name).strip(),
                                    "" if unit is None else str(unit).strip()))
            per_store.setdefault(u, {})[store_code] = (_num(qty), _num(sales))
    wb.close()
    return by_store, upc_info, per_store, unknown_stores


# ----------------------- populate the 4 store sheets ------------------------
def write_store_sheets(wb, by_store, add_totals=True):
    from openpyxl.styles import Font, PatternFill
    for code, items in by_store.items():
        ws = wb[code]
        for r in range(6, ws.max_row + 1):        # clear old data, keep headers 1-5
            for c in range(1, 13):
                ws.cell(row=r, column=c).value = None
        r = 6
        ws.cell(row=r, column=2).value = SHEET_DISPLAY[code]   # store name in B
        r += 1
        first_item_row = r
        seq, last_cat = 0, None
        for it in items:
            if it["cat"] != last_cat:
                ws.cell(row=r, column=3).value = it["cat"]      # category in C
                last_cat = it["cat"]; r += 1
            seq += 1
            ws.cell(row=r, column=4).value = seq
            ws.cell(row=r, column=5).value = it["name"]
            ws.cell(row=r, column=6).value = it["upc"]          # F = lookup key
            ws.cell(row=r, column=7).value = it["unit"]
            ws.cell(row=r, column=8).value = it["qty"]          # H = qty
            ws.cell(row=r, column=9).value = it["sales"]        # I = $
            r += 1
        if add_totals and r > first_item_row:
            last_item_row = r - 1
            tot = r + 1                                          # one blank spacer
            bold = Font(bold=True)
            fill = PatternFill("solid", fgColor="E9EEF5")
            ws.cell(row=tot, column=5, value="TOTAL").font = bold
            qc = ws.cell(row=tot, column=8, value="=SUM(H%d:H%d)" % (first_item_row, last_item_row))
            sc = ws.cell(row=tot, column=9, value="=SUM(I%d:I%d)" % (first_item_row, last_item_row))
            qc.font = bold; sc.font = bold
            sc.number_format = "#,##0.00"; qc.number_format = "#,##0"
            for c in range(5, 10):
                ws.cell(row=tot, column=c).fill = fill


# ----------------- read Summary master items (main block only) --------------
def read_summary_items(ws):
    """upc -> (category, name).  Reads only the main item block (stops before the
    embedded reconciliation/Table1 area) and ignores the pivot's tiny integers."""
    out = {}
    for r in range(FIRST_LEFT_ROW, SECONDARY_TABLE_START):
        B = ws.cell(row=r, column=2).value
        D = ws.cell(row=r, column=4).value
        E = ws.cell(row=r, column=5).value
        if B and str(B).strip() == "Total":
            continue
        if _is_upc(E):
            cat = str(B).strip() if B else "?"
            out[str(E).strip()] = (cat, "" if D is None else str(D).strip())
    return out


# -------------------- safe in-place VLOOKUP range widening -------------------
def widen_vlookup_ranges(ws):
    import re
    pat = re.compile(r"('00\d\d'!\$F\$1:\$I)\$\d+")   # e.g. '0008'!$F$1:$I$758
    fixed = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "VLOOKUP" in v:
                nv = pat.sub(lambda m: m.group(1).replace("$F$1:$I", "$F:$I"), v)
                nv = nv.replace("'0025'!$F:$II", "'0025'!$F:$I")
                if nv != v:
                    cell.value = nv; fixed += 1
    return fixed


# ----------------------- 12-month Trend table update ------------------------
def read_trend_frozen(template_path):
    """From LAST month's workbook, read the *computed* value of the newest Trend
    column (AE) for each category row. That cached value is last month's real
    number and is what we freeze into place before sliding the window."""
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb[SUMMARY_SHEET]
    frozen = {}
    for r in range(TREND_FIRST_ROW, TREND_LAST_ROW + 1):
        frozen[r] = ws.cell(row=r, column=TREND_LAST_COL).value
    wb.close()
    return frozen


def update_trend(ws, frozen, new_label):
    """Slide the 12-month window left by one month and add the new month.

    Before:  T..AC = 11 prior months | AD = 2nd-newest | AE = newest (formula)
    After :  T..AC <- U..AD (shift)   | AD <- frozen newest value | AE = formula
             (AE keeps its live formula, so it recomputes to `new_label`'s data
             once Excel recalculates the Summary from this month's store sheets.)

    Idempotent guard: if the newest header already equals `new_label`, the sheet
    was already advanced for this month -> skip so a re-run can't double-shift.
    """
    cur_newest = ws.cell(row=TREND_HDR_ROW, column=TREND_LAST_COL).value
    if cur_newest is not None and str(cur_newest).strip() == str(new_label).strip():
        print("Trend: already at '%s' -> skipped (idempotent)." % new_label)
        return 0

    # Snapshot the AE formula per row so the shift never disturbs it.
    ae_formula = {r: ws.cell(row=r, column=TREND_LAST_COL).value
                  for r in range(TREND_FIRST_ROW, TREND_LAST_ROW + 1)}

    # 1) Shift each category row: cols T..AC take the value from one col right;
    #    AD takes the frozen (computed) value of last month's AE.
    for r in range(TREND_FIRST_ROW, TREND_LAST_ROW + 1):
        for c in range(TREND_FIRST_COL, TREND_LAST_COL - 1):        # T..AC
            ws.cell(row=r, column=c).value = ws.cell(row=r, column=c + 1).value
        ws.cell(row=r, column=TREND_LAST_COL - 1).value = frozen.get(r)  # AD
        ws.cell(row=r, column=TREND_LAST_COL).value = ae_formula[r]      # AE (formula, unchanged)

    # 2) Shift the month headers the same way; AE header becomes the new month.
    for c in range(TREND_FIRST_COL, TREND_LAST_COL):                # T..AD
        ws.cell(row=TREND_HDR_ROW, column=c).value = ws.cell(row=TREND_HDR_ROW, column=c + 1).value
    ws.cell(row=TREND_HDR_ROW, column=TREND_LAST_COL).value = new_label

    # 3) Keep every month's column total consistent over ALL category rows
    #    (6..33). The template summed only 6..32 for older months, which would
    #    silently drop a category once it left the newest column; :33 is safe
    #    because empty history cells count as 0.
    from openpyxl.utils import get_column_letter
    for c in range(TREND_FIRST_COL, TREND_LAST_COL + 1):
        L = get_column_letter(c)
        ws.cell(row=TREND_TOTAL_ROW, column=c).value = \
            "=SUM(%s%d:%s%d)" % (L, TREND_FIRST_ROW, L, TREND_LAST_ROW)
    return 1


# --------------------------- change-log sheet -------------------------------
def write_change_log(wb, added, missing, unmapped, unknown_stores, month_label, notes=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    name = "Change Log"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    H = Font(bold=True, size=12); TH = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    r = 1
    def head(txt):
        nonlocal r
        c = ws.cell(row=r, column=1, value=txt); c.font = H; r += 2
    def tbl(cols, rows):
        nonlocal r
        for i, t in enumerate(cols, 1):
            c = ws.cell(row=r, column=i, value=t); c.font = TH; c.fill = fill
        r += 1
        for rec in rows:
            for i, v in enumerate(rec, 1):
                ws.cell(row=r, column=i, value=v)
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="AUTOMATED UPDATE - CHANGE LOG  (%s)" % month_label).font = Font(bold=True, size=14)
    r += 2
    if notes:
        head("DATA-INTEGRITY FIXES  (%d)" % len(notes))
        for n in notes:
            ws.cell(row=r, column=1, value="- " + n); r += 1
        r += 1
    # Only surface a NEW-ITEMS note when items actually appeared this month.
    if added:
        verb = "AUTO-INSERTED INTO THE SUMMARY" if AUTO_INSERT_ITEMS else "TO ADD"
        head("NEW ITEMS %s  (%d)  - sold this month, not in last month's Summary" % (verb, len(added)))
        if not AUTO_INSERT_ITEMS:
            ws.cell(row=r, column=1,
                    value="Insert a row in the category shown, put UPC in col E, then copy the G:P "
                          "formulas from the row above (Excel fixes Table1 automatically).")
            r += 2
        tbl(["Summary Category", "Item", "UPC", "Unit", "Total Qty", "Total $"],
            [(a["cat"], a["name"], a["upc"], a["unit"], a["qty"], a["sales"]) for a in added])
    else:
        head("NEW ITEMS  (0)")
        ws.cell(row=r, column=1, value="No new items in this month's Product Group Ranking.")
        r += 2

    head("NO-MOVEMENT / MISSING ITEMS  (%d)  - in Summary but no sales this month" % len(missing))
    tbl(["Category", "Item", "UPC"], [(m["cat"], m["name"], m["upc"]) for m in missing])

    if unmapped:
        head("!! COULD NOT MAP THESE CATEGORIES - decide placement manually  (%d)" % len(unmapped))
        tbl(["Raw Category", "Item", "UPC"], [(u["cat"], u["name"], u["upc"]) for u in unmapped])
    if unknown_stores:
        head("!! UNKNOWN STORE NAMES IN RAW FILE  (%d)" % len(unknown_stores))
        for s in sorted(unknown_stores):
            ws.cell(row=r, column=1, value=s); r += 1
    for col, w in {"A": 24, "B": 44, "C": 20, "D": 14, "E": 12, "F": 12}.items():
        ws.column_dimensions[col].width = w


# ---------------- fix a latent template bug in the grand-total $ ------------
def fix_grand_total_formula(ws):
    """The company 'Total DMK' $ cell (P392) sums the 28 category $-subtotals,
    but in the template its formula is MISSING the last term (+P387, World Cup
    Slushies) that every other column (H/J/L/N/O) includes -> the grand-total
    dollars are understated by that category every month. Rebuild P392 from the
    qty-total cell O392 (same row set) so the two can never diverge. Returns a
    human-readable note if a change was made, else None."""
    import re
    o392 = ws["O392"].value
    p392 = ws["P392"].value
    if not (isinstance(o392, str) and o392.startswith("=")):
        return None
    rebuilt = re.sub(r"\bO(\d+)", r"P\1", o392)   # O11+O21+... -> P11+P21+...
    if isinstance(p392, str) and rebuilt.replace(" ", "") == p392.replace(" ", ""):
        return None                                # already consistent
    ws["P392"].value = rebuilt
    return ("Fixed grand-total $ (P392): its SUM was missing a category term "
            "present in every other column, understating monthly totals. "
            "Rebuilt to match the units-total row set.")


# ------------------------------- housekeeping -------------------------------
def clean_junk_sheets(wb):
    """Drop the empty placeholder 'Sheet1' (held only a concept image) and the
    bloated 'No Movement old' (16,378 phantom columns) that inflate the file."""
    removed = []
    for name in ("Sheet1", "No Movement old"):
        if name in wb.sheetnames:
            del wb[name]
            removed.append(name)
    return removed


# ---------- keep the ORIGINAL Summary charts looking like the template ------
def restyle_original_charts(ws):
    """openpyxl preserves the template charts' teal chart-area fill, red line and
    positions on save, but it drops their (cell-linked) TITLES and the plot-area
    fill - which is what made them look wrong. Restore both so the Summary charts
    match the original June look: white bold category titles + teal plot area.
    Also drop the one broken (#REF!) placeholder chart. Returns (kept, titled)."""
    import re
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.text import CharacterProperties, Font as DFont
    TEAL = "097A95"                       # the template's chart-area teal
    keep, titled = [], 0
    for c in ws._charts:
        f = None
        try:
            f = c.series[0].val.numRef.f
        except Exception:
            pass
        if f and "#REF!" in f:            # broken placeholder -> discard
            continue
        # Restyle ONLY the small single-row per-category trend charts (those are
        # the ones openpyxl strips of title + plot fill). Every other chart -
        # notably the big multi-series "Catogory Chart" - is preserved perfectly
        # by openpyxl, so leave it completely untouched.
        rows = list({m for m in re.findall(r"\$[A-Z]+\$(\d+)", f)}) if f else []
        is_single_cat = (len(rows) == 1 and (ws.cell(int(rows[0]), 19).value is not None))
        if not is_single_cat:
            # The big multi-series "Catogory Chart" keeps its title, gradient,
            # legend and series colours through openpyxl - but its DATA TABLE
            # (the value grid under the lines) is dropped. Restore it.
            if f and "$D$" in f:
                from openpyxl.chart.plotarea import DataTable
                c.plot_area.dTable = DataTable(showHorzBorder=True, showVertBorder=True,
                                               showOutline=True, showKeys=True)
            keep.append(c)
            continue
        cat = str(ws.cell(int(rows[0]), 19).value)
        c.plot_area.graphicalProperties = GraphicalProperties(solidFill=TEAL)
        c.title = cat
        try:                              # white + bold title like the original
            run = c.title.tx.rich.p[0].r[0]
            run.rPr = CharacterProperties(sz=1000, b=True, solidFill="FFFFFF",
                                          latin=DFont(typeface="Calibri"))
        except Exception:
            pass
        titled += 1
        keep.append(c)
    ws._charts = keep
    return len(keep), titled


# -------------- rebuild the Summary per-category charts (clean) -------------
def rebuild_summary_charts(ws):
    """Delete the degraded per-category line charts (openpyxl can't preserve
    their original styling on save) and rebuild them uniformly: one tidy
    sparkline-style 12-month trend per category, in a grid to the right of the
    Trend table, plus one clean category-ranking bar chart below Table1."""
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.axis import ChartLines
    n_removed = len(getattr(ws, "_charts", []))
    ws._charts = []            # drop every existing chart on the sheet

    months = Reference(ws, min_col=TREND_FIRST_COL, max_col=TREND_LAST_COL,
                       min_row=TREND_HDR_ROW, max_row=TREND_HDR_ROW)

    # ---- per-category mini trend charts, 4 across x 7 down --------------
    cols = [33, 41, 49, 57]           # AG, AO, AW, BE
    row_step, row0 = 12, 36
    built = 0
    for idx, trow in enumerate(range(TREND_FIRST_ROW, TREND_LAST_ROW + 1)):
        cat = ws.cell(row=trow, column=19).value        # S = category name
        if cat is None:
            continue
        ch = LineChart()
        ch.title = str(cat)
        ch.height, ch.width = 5.2, 8.2
        ch.style = 2
        dat = Reference(ws, min_col=TREND_FIRST_COL, max_col=TREND_LAST_COL,
                        min_row=trow, max_row=trow)
        ch.add_data(dat, from_rows=True, titles_from_data=False)
        ch.set_categories(months)
        s = ch.series[0]
        s.smooth = True
        s.graphicalProperties.line.solidFill = CLR_LSK
        s.graphicalProperties.line.width = 26000       # ~2pt
        ch.legend = None
        ch.x_axis.delete = True                        # sparkline: no axes clutter
        ch.y_axis.delete = True
        ch.x_axis.majorGridlines = None
        ch.y_axis.majorGridlines = None
        anchor = "%s%d" % (openpyxl.utils.get_column_letter(cols[idx % 4]),
                           row0 + (idx // 4) * row_step)
        ws.add_chart(ch, anchor)
        built += 1

    # ---- one clean category-ranking bar chart (below Table1) -----------
    newest = ws.cell(row=TREND_HDR_ROW, column=TREND_LAST_COL).value or ""
    bar = BarChart(); bar.type = "bar"
    bar.title = "Category Volume by Month  (%s, units)" % str(newest).strip()
    bar.height, bar.width = 16, 22
    bdat = Reference(ws, min_col=TREND_LAST_COL, min_row=TREND_FIRST_ROW, max_row=TREND_LAST_ROW)
    bcat = Reference(ws, min_col=19, min_row=TREND_FIRST_ROW, max_row=TREND_LAST_ROW)
    bar.add_data(bdat, titles_from_data=False); bar.set_categories(bcat)
    bar.series[0].graphicalProperties.solidFill = CLR_LSK
    bar.legend = None
    bar.x_axis.majorGridlines = ChartLines()
    bar.gapWidth = 40
    bar.x_axis.delete = False; bar.y_axis.delete = False
    ws.add_chart(bar, "B437")
    return n_removed, built


# ------------------------------ dashboard sheet -----------------------------
def build_dashboard(wb, period_label):
    """Create a 'Dashboard' summary sheet: KPI tiles + professional charts, all
    driven by live formula references into the Summary reconciliation block
    (rows 390-394) and the Trend totals, so Excel repairs them on recalc/insert."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import Marker

    def val_labels(numfmt=None):
        """Value-only data labels (no series/category-name clutter)."""
        dl = DataLabelList()
        dl.showVal = True
        dl.showSerName = False
        dl.showCatName = False
        dl.showLegendKey = False
        dl.showPercent = False
        dl.showBubbleSize = False
        if numfmt:
            dl.numFmt = numfmt
        return dl

    S = "'Summary Comparision Report'"
    name = "Dashboard"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    ws.sheet_view.showGridLines = False

    ink   = Font(name="Calibri", color=CLR_INK)
    mute  = Font(name="Calibri", color=CLR_MUTE, size=9)
    title = Font(name="Calibri Light", size=22, bold=True, color=CLR_INK)
    sub   = Font(name="Calibri", size=12, color=CLR_MUTE)

    # ---- Banner ----
    ws["B2"] = "MTO & G&G  —  FMO Report"
    ws["B2"].font = title
    ws["B3"] = "Foodservice sales · 4 Demo Market c-stores · %s" % period_label
    ws["B3"].font = sub

    # ---- KPI tiles (live formulas off the reconciliation block) ----
    tiles = [
        ("Grand Total Sales",   "=%s!$P$394" % S, "$#,##0",  CLR_COMPD),
        ("Total DMK Sales",     "=%s!$P$392" % S, "$#,##0",  CLR_LSK),
        ("Uber & GrubHub",      "=%s!$P$393" % S, "$#,##0",  CLR_UBER),
        ("Total Units (DMK)",   "=%s!$O$392" % S, "#,##0",   CLR_INK),
        # qty month-over-month from the Trend totals (AD=prev, AE=this month)
        ("Units vs Last Month",
         "=IFERROR((%s!$AE$34-%s!$AD$34)/%s!$AD$34,0)" % (S, S, S), "0.0%", CLR_INK),
    ]
    thin = Side(style="thin", color="D9DEE6")
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)
    col = 2
    for label, formula, numfmt, accent in tiles:
        # accent tile spanning 2 columns; value cell merged so big $ fit.
        c1, c2 = col, col + 1
        for rr in (5, 6, 7):
            for cc in (c1, c2):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = PatternFill("solid", fgColor=CLR_BAND)
                cell.border = box
        for rr in (5, 6, 7):
            ws.merge_cells(start_row=rr, start_column=c1, end_row=rr, end_column=c2)
        lc = ws.cell(row=5, column=c1, value=label)
        lc.font = Font(name="Calibri", size=9, bold=True, color=CLR_MUTE)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=6, column=c1, value=formula)
        vc.font = Font(name="Calibri", size=16, bold=True, color=accent)
        vc.number_format = numfmt
        vc.alignment = Alignment(horizontal="left", vertical="center")
        pc = ws.cell(row=7, column=c1, value=period_label); pc.font = mute
        pc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[6].height = 24
        col += 2

    # ---- Staging table for charts (hidden columns AA+) -------------------
    # Per-store $ split: Total DMK / Uber / Compound  (cols H,J,L,N of recon rows)
    stores = [("Northgate", "H"), ("Riverside", "J"), ("Lakeview", "L"), ("Summit", "N")]
    hdr = ["Store", "Total DMK", "Uber & GrubHub", "Compound"]
    base_r = 40           # staging starts well below the visual area
    sc = 27               # column AA
    for j, h in enumerate(hdr):
        ws.cell(row=base_r, column=sc + j, value=h).font = Font(bold=True)
    for i, (store, colL) in enumerate(stores, 1):
        rr = base_r + i
        ws.cell(row=rr, column=sc + 0, value=store)
        ws.cell(row=rr, column=sc + 1, value="=%s!$%s$392" % (S, colL)).number_format = "$#,##0"
        ws.cell(row=rr, column=sc + 2, value="=%s!$%s$393" % (S, colL)).number_format = "$#,##0"
        ws.cell(row=rr, column=sc + 3, value="=%s!$%s$394" % (S, colL)).number_format = "$#,##0"

    # MTO vs G&G per store ($) staging
    m_r = base_r + 7
    ws.cell(row=m_r, column=sc + 0, value="Store").font = Font(bold=True)
    ws.cell(row=m_r, column=sc + 1, value="MTO").font = Font(bold=True)
    ws.cell(row=m_r, column=sc + 2, value="G&G").font = Font(bold=True)
    for i, (store, colL) in enumerate(stores, 1):
        rr = m_r + i
        ws.cell(row=rr, column=sc + 0, value=store)
        ws.cell(row=rr, column=sc + 1, value="=%s!$%s$390" % (S, colL)).number_format = "$#,##0"
        ws.cell(row=rr, column=sc + 2, value="=%s!$%s$391" % (S, colL)).number_format = "$#,##0"

    def _fill(series, hexcolor):
        series.graphicalProperties.solidFill = hexcolor
        series.graphicalProperties.line.solidFill = hexcolor

    # ---- Chart 1: per-store sales split (clustered columns) --------------
    ch1 = BarChart(); ch1.type = "col"; ch1.grouping = "clustered"
    ch1.title = "Sales by Store  —  DMK / Delivery / Compound"
    ch1.height = 8.5; ch1.width = 17
    data = Reference(ws, min_col=sc + 1, max_col=sc + 3, min_row=base_r, max_row=base_r + 4)
    cats = Reference(ws, min_col=sc + 0, min_row=base_r + 1, max_row=base_r + 4)
    ch1.add_data(data, titles_from_data=True); ch1.set_categories(cats)
    for s, clr in zip(ch1.series, (CLR_LSK, CLR_UBER, CLR_COMPD)):
        _fill(s, clr)
    ch1.dataLabels = val_labels("$#,##0"); ch1.gapWidth = 60
    ch1.y_axis.numFmt = "$#,##0"; ch1.y_axis.majorGridlines = None
    ch1.x_axis.delete = False; ch1.y_axis.delete = False
    ch1.legend.position = "b"          # legend below the plot, off the bars
    ws.add_chart(ch1, "B9")

    # ---- Chart 2: 12-month company volume trend (line) -------------------
    ch2 = LineChart()
    ch2.title = "Company Volume — 12-Month Trend (units)"
    ch2.height = 8.5; ch2.width = 17
    # NOTE: the trend row lives on the Summary sheet; pull it cross-sheet.
    tref = Reference(wb["Summary Comparision Report"], min_col=20, max_col=31, min_row=34, max_row=34)
    tcat = Reference(wb["Summary Comparision Report"], min_col=20, max_col=31, min_row=5, max_row=5)
    ch2.add_data(tref, from_rows=True, titles_from_data=False)
    ch2.set_categories(tcat)
    ch2.series[0].smooth = False
    ch2.series[0].graphicalProperties.line.solidFill = CLR_TREND
    ch2.series[0].graphicalProperties.line.width = 28000   # ~2.2pt
    ch2.series[0].marker = Marker(symbol="circle", size=6)
    ch2.series[0].marker.graphicalProperties.solidFill = CLR_TREND
    ch2.dataLabels = None            # no per-point clutter; axis carries the months
    ch2.legend = None
    ch2.x_axis.delete = False; ch2.y_axis.delete = False
    ch2.y_axis.numFmt = "#,##0"; ch2.y_axis.majorGridlines = None
    ch2.y_axis.scaling.min = 15000   # lift the floor so month-to-month shape is visible
    ws.add_chart(ch2, "J9")

    # ---- Chart 3: MTO vs G&G by store (clustered columns) ----------------
    ch3 = BarChart(); ch3.type = "col"; ch3.grouping = "clustered"
    ch3.title = "MTO vs G&G by Store ($)"
    ch3.height = 8.5; ch3.width = 17
    d3 = Reference(ws, min_col=sc + 1, max_col=sc + 2, min_row=m_r, max_row=m_r + 4)
    c3 = Reference(ws, min_col=sc + 0, min_row=m_r + 1, max_row=m_r + 4)
    ch3.add_data(d3, titles_from_data=True); ch3.set_categories(c3)
    for s, clr in zip(ch3.series, (CLR_MTO, CLR_GG)):
        _fill(s, clr)
    ch3.dataLabels = val_labels("$#,##0"); ch3.gapWidth = 60
    ch3.y_axis.numFmt = "$#,##0"; ch3.y_axis.majorGridlines = None
    ch3.x_axis.delete = False; ch3.y_axis.delete = False
    ch3.legend.position = "b"
    ws.add_chart(ch3, "B26")

    # ---- Chart 4: category volume this month (bar) -----------------------
    ch4 = BarChart(); ch4.type = "bar"
    ch4.title = "Category Volume — %s (units)" % period_label
    ch4.height = 12; ch4.width = 17
    sumws = wb["Summary Comparision Report"]
    d4 = Reference(sumws, min_col=31, min_row=6, max_row=33)   # AE6:AE33 (newest month qty)
    c4 = Reference(sumws, min_col=19, min_row=6, max_row=33)   # S6:S33 category names
    ch4.add_data(d4, titles_from_data=False); ch4.set_categories(c4)
    _fill(ch4.series[0], CLR_LSK)
    ch4.legend = None; ch4.x_axis.majorGridlines = None
    ch4.x_axis.delete = False; ch4.y_axis.delete = False
    ch4.gapWidth = 40
    ws.add_chart(ch4, "J26")

    # column widths for the KPI band
    for cc in range(2, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(cc)].width = 13
    # print only the visual block; keep the staging tables (cols AA+) off the page
    ws.print_area = "A1:T52"
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return ws


# --------------------- standalone interactive HTML dashboard ----------------
def build_html_dashboard(out_xlsx, html_path, period_label):
    """Read the recalculated workbook and emit a self-contained (offline, no
    external libraries) HTML dashboard with inline-SVG charts. Safe to email /
    open in any browser. Reconciliation rows are found by LABEL because the COM
    insert may have shifted them."""
    wb = openpyxl.load_workbook(out_xlsx, data_only=True)
    ws = wb[SUMMARY_SHEET]

    lbl = {}
    for r in range(384, 410):
        f = ws.cell(r, 6).value
        if f:
            lbl[str(f).strip()] = r
    r_mto  = lbl.get("MTO"); r_gg = lbl.get("G&G"); r_lsk = lbl.get("Total DMK")
    r_uber = lbl.get("Add: Uber & Grub Hub sales"); r_comp = lbl.get("Total DMK and U&GH")

    STORES = [("Northgate", 8), ("Riverside", 10), ("Lakeview", 12), ("Summit", 14)]
    def num(v): return float(v) if isinstance(v, (int, float)) else 0.0
    def row_by_store(r): return {s: num(ws.cell(r, c).value) for s, c in STORES} if r else {}

    lsk, uber, comp = row_by_store(r_lsk), row_by_store(r_uber), row_by_store(r_comp)
    mto, gg = row_by_store(r_mto), row_by_store(r_gg)
    grand   = num(ws.cell(r_comp, 16).value) if r_comp else 0.0
    tot_lsk = num(ws.cell(r_lsk, 16).value) if r_lsk else 0.0
    tot_ube = num(ws.cell(r_uber, 16).value) if r_uber else 0.0
    units   = num(ws.cell(r_lsk, 15).value) if r_lsk else 0.0

    months = [ws.cell(TREND_HDR_ROW, c).value for c in range(TREND_FIRST_COL, TREND_LAST_COL + 1)]
    trend  = [num(ws.cell(TREND_TOTAL_ROW, c).value) for c in range(TREND_FIRST_COL, TREND_LAST_COL + 1)]
    prev_units = trend[-2] if len(trend) >= 2 else 0.0
    mom = (units - prev_units) / prev_units if prev_units else 0.0

    cats = []
    for r in range(TREND_FIRST_ROW, TREND_LAST_ROW + 1):
        nm = ws.cell(r, 19).value
        if nm is not None:
            cats.append((str(nm), num(ws.cell(r, TREND_LAST_COL).value)))
    cats.sort(key=lambda x: x[1], reverse=True)
    wb.close()

    C = dict(lsk="#%s" % CLR_LSK, uber="#%s" % CLR_UBER, comp="#%s" % CLR_COMPD,
             gg="#%s" % CLR_GG, ink="#%s" % CLR_INK, mute="#%s" % CLR_MUTE)

    def esc(s): return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    def money(v): return "$%s" % format(int(round(v)), ",")
    def qty(v):   return format(int(round(v)), ",")

    # ---- grouped vertical bars (stores x series) -----------------------
    def grouped_bars(data_by_series, series_colors, val_fmt, w=560, h=320):
        pad_l, pad_b, pad_t, pad_r = 56, 44, 20, 12
        stores = [s for s, _ in STORES]
        allv = [v for d in data_by_series.values() for v in d.values()] or [0]
        vmax = max(allv) * 1.15 or 1
        pw, ph = w - pad_l - pad_r, h - pad_t - pad_b
        gw = pw / len(stores)
        ns = len(data_by_series)
        bw = gw * 0.8 / ns
        out = ['<svg viewBox="0 0 %d %d" class="chart">' % (w, h)]
        for gl_ in range(5):
            y = pad_t + ph - ph * gl_ / 4
            val = vmax * gl_ / 4
            out.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" class="grid"/>' % (pad_l, y, w - pad_r, y))
            out.append('<text x="%d" y="%.1f" class="ax ar">%s</text>' % (pad_l - 6, y + 4, val_fmt(val)))
        for gi, store in enumerate(stores):
            gx = pad_l + gi * gw + gw * 0.1
            for si, (sname, d) in enumerate(data_by_series.items()):
                v = d.get(store, 0)
                bh = ph * v / vmax
                x = gx + si * bw
                y = pad_t + ph - bh
                out.append('<rect x="%.1f" y="%.1f" width="%.1f" height="%.1f" rx="2" fill="%s">'
                           '<title>%s — %s: %s</title></rect>'
                           % (x, y, bw - 2, bh, series_colors[si], store, esc(sname), val_fmt(v)))
                out.append('<text x="%.1f" y="%.1f" class="dl">%s</text>'
                           % (x + (bw - 2) / 2, y - 4, val_fmt(v)))
            out.append('<text x="%.1f" y="%d" class="ax am">%s</text>'
                       % (pad_l + gi * gw + gw / 2, h - pad_b + 18, store))
        out.append('</svg>')
        return "".join(out)

    # ---- line chart ----------------------------------------------------
    def line_chart(labels, values, color, w=560, h=320):
        pad_l, pad_b, pad_t, pad_r = 56, 46, 20, 14
        pw, ph = w - pad_l - pad_r, h - pad_t - pad_b
        vals = [v for v in values if v]
        lo = min(vals) * 0.9 if vals else 0
        hi = max(vals) * 1.08 if vals else 1
        rng = (hi - lo) or 1
        n = len(values)
        xs = [pad_l + (pw * i / (n - 1) if n > 1 else 0) for i in range(n)]
        ys = [pad_t + ph - ph * (v - lo) / rng for v in values]
        out = ['<svg viewBox="0 0 %d %d" class="chart">' % (w, h)]
        for gl_ in range(5):
            y = pad_t + ph - ph * gl_ / 4
            out.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" class="grid"/>' % (pad_l, y, w - pad_r, y))
            out.append('<text x="%d" y="%.1f" class="ax ar">%s</text>' % (pad_l - 6, y + 4, qty(lo + rng * gl_ / 4)))
        pts = " ".join("%.1f,%.1f" % (x, y) for x, y in zip(xs, ys))
        out.append('<polyline points="%s" fill="none" stroke="%s" stroke-width="2.5" stroke-linejoin="round"/>' % (pts, color))
        for i, (x, y) in enumerate(zip(xs, ys)):
            out.append('<circle cx="%.1f" cy="%.1f" r="3.5" fill="%s"><title>%s: %s</title></circle>'
                       % (x, y, color, esc(labels[i]), qty(values[i])))
            out.append('<text x="%.1f" y="%d" class="ax am" transform="rotate(-40 %.1f %d)">%s</text>'
                       % (x, h - pad_b + 16, x, h - pad_b + 16, esc(labels[i])))
        out.append('</svg>')
        return "".join(out)

    # ---- horizontal bars (categories) ----------------------------------
    def hbars(pairs, color, w=560, rowh=20):
        pad_l, pad_r, pad_t = 150, 60, 10
        h = pad_t * 2 + rowh * len(pairs)
        vmax = max((v for _, v in pairs), default=1) or 1
        pw = w - pad_l - pad_r
        out = ['<svg viewBox="0 0 %d %d" class="chart">' % (w, h)]
        for i, (name, v) in enumerate(pairs):
            y = pad_t + i * rowh
            bw = pw * v / vmax
            out.append('<text x="%d" y="%.1f" class="ax ar cl">%s</text>' % (pad_l - 6, y + rowh * 0.7, esc(name)))
            out.append('<rect x="%d" y="%.1f" width="%.1f" height="%.1f" rx="2" fill="%s"><title>%s: %s</title></rect>'
                       % (pad_l, y + 2, bw, rowh - 6, color, esc(name), qty(v)))
            out.append('<text x="%.1f" y="%.1f" class="dl dll">%s</text>' % (pad_l + bw + 4, y + rowh * 0.7, qty(v)))
        out.append('</svg>')
        return "".join(out)

    def legend(items):
        sp = "".join('<span class="lg"><i style="background:%s"></i>%s</span>' % (c, esc(t)) for t, c in items)
        return '<div class="legend">%s</div>' % sp

    kpis = [
        ("Grand Total Sales", money(grand), C["comp"]),
        ("Total DMK Sales", money(tot_lsk), C["lsk"]),
        ("Uber & GrubHub", money(tot_ube), C["uber"]),
        ("Total Units (DMK)", qty(units), C["ink"]),
        ("Units vs Last Month", "%+.1f%%" % (mom * 100), C["comp"] if mom >= 0 else "#C0392B"),
    ]
    kpi_html = "".join(
        '<div class="kpi"><div class="kl">%s</div><div class="kv" style="color:%s">%s</div>'
        '<div class="kp">%s</div></div>' % (esc(k), c, esc(v), esc(period_label))
        for k, v, c in kpis)

    sales = grouped_bars(OrderedDict([("Total DMK", lsk), ("Uber & GrubHub", uber), ("Compound", comp)]),
                         [C["lsk"], C["uber"], C["comp"]], money)
    trendc = line_chart([str(m) for m in months], trend, C["comp"])
    mtogg = grouped_bars(OrderedDict([("MTO", mto), ("G&G", gg)]), [C["lsk"], C["gg"]], money)
    catc  = hbars(cats, C["lsk"])

    html = """<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>MTO &amp; G&amp;G — FMO %(period)s</title>
<style>
 :root{--ink:%(ink)s;--mute:%(mute)s;--line:#E3E8EF;--card:#fff;--bg:#F4F6FA}
 *{box-sizing:border-box}
 body{margin:0;background:var(--bg);color:var(--ink);
   font-family:'Segoe UI',Roboto,Helvetica,Arial,sans-serif;-webkit-font-smoothing:antialiased}
 .wrap{max-width:1200px;margin:0 auto;padding:28px 22px 60px}
 h1{font-size:26px;font-weight:700;margin:0}
 .sub{color:var(--mute);font-size:14px;margin:4px 0 22px}
 .kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:22px}
 .kpi{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 16px;
   box-shadow:0 1px 2px rgba(16,24,40,.04)}
 .kl{font-size:11px;font-weight:600;letter-spacing:.02em;color:var(--mute);text-transform:uppercase}
 .kv{font-size:26px;font-weight:700;margin-top:6px;line-height:1.1}
 .kp{font-size:11px;color:var(--mute);margin-top:4px}
 .grid2{display:grid;grid-template-columns:1fr 1fr;gap:18px}
 .card{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:16px 18px;
   box-shadow:0 1px 2px rgba(16,24,40,.04)}
 .card h2{font-size:15px;margin:0 0 4px}
 .card.full{grid-column:1/-1}
 .chart{width:100%%;height:auto;display:block;overflow:visible}
 .grid{stroke:#EDF1F6;stroke-width:1}
 .ax{fill:var(--mute);font-size:11px}
 .ar{text-anchor:end}.am{text-anchor:middle}.cl{text-anchor:end;dominant-baseline:middle}
 .dl{fill:var(--ink);font-size:10.5px;font-weight:600;text-anchor:middle}
 .dll{text-anchor:start;dominant-baseline:middle;fill:var(--mute)}
 .legend{display:flex;gap:16px;flex-wrap:wrap;margin-top:10px;justify-content:center}
 .lg{font-size:12px;color:var(--mute);display:inline-flex;align-items:center;gap:6px}
 .lg i{width:11px;height:11px;border-radius:3px;display:inline-block}
 rect{transition:opacity .12s}rect:hover{opacity:.82}
 .foot{color:var(--mute);font-size:11px;margin-top:26px;text-align:center}
 @media(max-width:860px){.kpis{grid-template-columns:repeat(2,1fr)}.grid2{grid-template-columns:1fr}}
</style></head><body><div class="wrap">
 <h1>MTO &amp; G&amp;G — FMO Report</h1>
 <div class="sub">Foodservice sales · 4 Demo Market c-stores · %(period)s</div>
 <div class="kpis">%(kpis)s</div>
 <div class="grid2">
   <div class="card"><h2>Sales by Store</h2>%(sales)s%(saleslg)s</div>
   <div class="card"><h2>Company Volume — 12-Month Trend (units)</h2>%(trend)s</div>
   <div class="card"><h2>MTO vs G&amp;G by Store ($)</h2>%(mtogg)s%(mtoglg)s</div>
   <div class="card"><h2>Category Volume (units)</h2>%(cat)s</div>
 </div>
 <div class="foot">Generated automatically from the monthly workbook · figures reconcile to the Summary tab.</div>
</div></body></html>""" % dict(
        period=esc(period_label), ink=C["ink"], mute=C["mute"], kpis=kpi_html,
        sales=sales, saleslg=legend([("Total DMK", C["lsk"]), ("Uber & GrubHub", C["uber"]), ("Compound", C["comp"])]),
        trend=trendc, mtogg=mtogg, mtoglg=legend([("MTO", C["lsk"]), ("G&G", C["gg"])]), cat=catc)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    return html_path


# ----------------- hands-off finish: Excel COM (insert + recalc) ------------
def com_finalize(out_file, added, do_insert, do_recalc, template_file=None, copy_charts=False):
    """Open the finished workbook in Excel via COM. Optionally copy the Summary
    charts across from the pristine template (Excel-to-Excel, so their original
    styling is kept exactly - the Python library can't preserve it). Insert each
    NEW item into its Summary category (Excel self-repairs Table1 / reconciliation
    / trend and every dependent reference), then full-recalc and save.
    Returns (inserted, ok)."""
    import win32com.client as win32
    import pythoncom
    S_SHEET = SUMMARY_SHEET
    xlDown = -4121
    inserted = 0
    pythoncom.CoInitialize()
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AskToUpdateLinks = False
    xl.ScreenUpdating = False
    try:
        wb = xl.Workbooks.Open(os.path.abspath(out_file), UpdateLinks=0)
        ws = wb.Worksheets(S_SHEET)

        # ---- restore the ORIGINAL Summary charts from the template ----
        if copy_charts and template_file:
            try:
                src_wb = xl.Workbooks.Open(os.path.abspath(template_file), ReadOnly=True, UpdateLinks=0)
                try:
                    src_ws = src_wb.Worksheets(S_SHEET)
                    n_src = src_ws.ChartObjects().Count
                    if ws.ChartObjects().Count:
                        ws.ChartObjects().Delete()      # drop the library-stripped copies
                    if n_src:
                        src_ws.Activate()
                        src_ws.ChartObjects().Copy()    # copy all, positions preserved
                        ws.Activate()
                        ws.Paste()
                        xl.CutCopyMode = False
                        # Copying between workbooks re-points each series to the
                        # SOURCE file (external link). Strip that token so the
                        # charts read THIS workbook's cells (the new month's data).
                        token = "[%s]" % os.path.basename(template_file)
                        for co in ws.ChartObjects():
                            try:
                                for sr in co.Chart.SeriesCollection():
                                    f = sr.Formula
                                    if token in f:
                                        sr.Formula = f.replace(token, "")
                            except Exception:
                                pass
                        print("   restored %d original Summary chart(s) from template." % n_src)
                finally:
                    src_wb.Close(SaveChanges=False)
            except Exception as ce:
                print("   !! chart restore skipped (%s)" % ce)

        if do_insert and added:
            # Group new items by their target category, insert per category.
            from collections import defaultdict
            by_cat = defaultdict(list)
            for a in added:
                by_cat[a["cat"]].append(a)

            def find_total_row(cat):
                """Row of the 'Total' subtotal that closes category `cat`
                (search col B: category header, then its following 'Total')."""
                last = ws.UsedRange.Row + ws.UsedRange.Rows.Count
                r = FIRST_LEFT_ROW
                in_cat = False
                while r < SECONDARY_TABLE_START:
                    b = ws.Cells(r, 2).Value
                    b = "" if b is None else str(b).strip()
                    if b == cat:
                        in_cat = True
                    elif in_cat and b == "Total":
                        return r
                    r += 1
                return None

            xlShiftDown = -4121
            for cat, items in by_cat.items():
                for it in items:
                    total_r = find_total_row(cat)
                    if not total_r:
                        print("  !! COM: category not found, skipped:", cat, it["upc"])
                        continue
                    src = total_r - 1                 # last existing item row
                    # Insert ONLY the item-block cells A:P (cols 1-16) and shift
                    # them down, so the side-by-side Trend table (cols R-AE) and
                    # the far-right charts are NOT disturbed. Excel still repairs
                    # Table1 / reconciliation / trend formula refs and expands the
                    # category subtotal SUM to include the new row.
                    rng = ws.Range(ws.Cells(src, 1), ws.Cells(src, 16))
                    rng.Copy()
                    rng.Insert(xlShiftDown)           # "insert copied cells", A:P only
                    xl.CutCopyMode = False
                    ws.Cells(src, 2).Value = cat      # B  category
                    ws.Cells(src, 4).Value = it["name"]   # D name
                    ws.Cells(src, 5).Value = "'" + it["upc"]  # E UPC as text
                    ws.Cells(src, 6).Value = it["unit"]   # F unit
                    inserted += 1

        if do_recalc:
            xl.CalculateFull()
        wb.Save()
        wb.Close(SaveChanges=True)
        return inserted, True
    finally:
        xl.Quit()
        pythoncom.CoUninitialize()


# --------------------------------- main -------------------------------------
def main():
    cfg = resolve_config()
    RAW, TEMPLATE, OUT = cfg["raw"], cfg["template"], cfg["out"]
    MONTH_LABEL, PERIOD = cfg["label"], cfg["period"]
    month_label = os.path.basename(OUT)
    print("Raw     :", os.path.basename(RAW))
    print("Template:", os.path.basename(TEMPLATE))
    print("Output  :", os.path.basename(OUT))
    print("New month: %s   (dashboard title: %s)" % (MONTH_LABEL, PERIOD))
    by_store, upc_info, per_store, unknown_stores = parse_raw(RAW)
    for code, items in by_store.items():
        tq = sum(i["qty"] for i in items); ts = sum(i["sales"] for i in items)
        print("  %s %-24s %4d items  qty %7.0f  $ %11.2f" %
              (code, SHEET_DISPLAY[code], len(items), tq, ts))
    if unknown_stores:
        print("  !! UNKNOWN stores:", unknown_stores)

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb[SUMMARY_SHEET]
    sum_items = read_summary_items(ws)
    print("Summary master items:", len(sum_items))

    raw_upcs = set(upc_info)
    new_upcs = sorted(raw_upcs - set(sum_items))
    missing_upcs = sorted(set(sum_items) - raw_upcs)

    added, unmapped = [], []
    summary_cats = {c for c, _ in sum_items.values()}
    for u in new_upcs:
        rawcat, name, unit = upc_info[u]
        target = CAT_MAP.get(rawcat, rawcat)
        tq = sum(v[0] for v in per_store[u].values())
        ts = sum(v[1] for v in per_store[u].values())
        rec = dict(cat=target, name=name, upc=u, unit=unit, qty=tq, sales=ts)
        (added if target in summary_cats else unmapped).append(
            rec if target in summary_cats else dict(cat=rawcat, name=name, upc=u))

    missing = [dict(cat=sum_items[u][0], name=sum_items[u][1], upc=u) for u in missing_upcs]

    print("NEW items to add   :", len(added))
    print("Unmapped categories:", len(unmapped))
    print("No-movement items  :", len(missing))

    if DRY_RUN:
        print("\nDRY_RUN=True -> nothing written.")
        return

    write_store_sheets(wb, by_store, add_totals=ADD_STORE_TOTALS)
    if FIX_VLOOKUP_RANGES:
        n = widen_vlookup_ranges(ws)
        print("VLOOKUP ranges widened:", n)
    if UPDATE_TREND:
        frozen = read_trend_frozen(TEMPLATE)
        if update_trend(ws, frozen, MONTH_LABEL):
            print("Trend: window advanced -> newest column = '%s'" % MONTH_LABEL)
    notes = []
    fix = fix_grand_total_formula(ws)
    if fix:
        notes.append(fix); print("Integrity:", fix)
    if REBUILD_SUMMARY_CHARTS:
        rem, built = rebuild_summary_charts(ws)
        print("Summary charts: removed %d degraded, rebuilt %d clean per-category + 1 ranking." % (rem, built))
    elif PRESERVE_SUMMARY_CHARTS:
        kept, titled = restyle_original_charts(ws)
        print("Summary charts: kept %d original-style, restored %d titles + teal fill (dropped broken)." % (kept, titled))
    if CLEAN_JUNK_SHEETS:
        rm = clean_junk_sheets(wb)
        print("Junk sheets removed:", rm or "(none)")
    if BUILD_DASHBOARD:
        build_dashboard(wb, PERIOD)
        print("Dashboard sheet built.")
    write_change_log(wb, added, missing, unmapped, unknown_stores, month_label, notes)
    wb.save(OUT)
    print("\nSaved (openpyxl stage):", os.path.basename(OUT))

    # Hands-off finish: insert new items + recalc + save (charts already handled).
    if AUTO_INSERT_ITEMS or RECALC_WITH_EXCEL:
        try:
            ins, ok = com_finalize(OUT, added,
                                   do_insert=AUTO_INSERT_ITEMS, do_recalc=RECALC_WITH_EXCEL)
            print("Excel finish: %d item(s) inserted, recalculated, saved." % ins)
        except Exception as e:
            print("!! Excel COM finish failed (%s). The openpyxl file is still valid;" % e)
            print("   open it in Excel, press F9, and paste any new items from the Change Log.")
    else:
        print("Open it in Excel and press F9 so the VLOOKUPs refresh.")

    # Standalone HTML dashboard (reads the recalculated values).
    if BUILD_HTML_DASHBOARD:
        try:
            html_path = os.path.join(BASE, "Dashboard %s.html" % PERIOD)
            build_html_dashboard(OUT, html_path, PERIOD)
            print("HTML dashboard:", os.path.basename(html_path))
        except Exception as e:
            print("!! HTML dashboard failed:", e)


if __name__ == "__main__":
    main()
